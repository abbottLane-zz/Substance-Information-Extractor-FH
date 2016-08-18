from DataLoading import ServerQuery
from SystemUtilities.Configuration import ADJUDICATION_DIR
from SystemUtilities.Globals import ALL_ATTRIBS
import openpyxl


def main():
    """ Create patients objects where values are those found in the adjudicated excel file and the spans are the
     union of spans across the annotators"""
    annotations = ServerQuery.get_annotations_from_server()

    # create adjudicated 'patients' object -- List[Patient]
    values = get_adjudicated_values()

    patients = {}
    first_annotator = annotations.keys()[0]

    for mrn, docs in annotations[first_annotator].values():
        if mrn not in patients:
            patients[mrn] = {}

        for doc_id, events in docs.values():
            if doc_id not in patients[mrn]:
                patients[mrn][doc_id] = events

            for event in events:
                substance = event.substance_type
                for attribute in ALL_ATTRIBS[substance]:
                    field_name = event.substance_type + attribute

                    if field_name in values[doc_id]:
                        event.status = values[doc_id][field_name]


    # save to pickle





def get_adjudicated_values():
    """
    Return values from excel file containing adjudicated batch 0 values
    :return: {doc_id: {field_name: value}}
    """
    wb = openpyxl.load_workbook(ADJUDICATION_DIR)
    sheets = wb.get_sheet_names()
    sheet_name = sheets[0]
    sheet = wb.get_sheet_by_name(sheet_name)

    num_of_rows = sheet.get_highest_row()
    doc_column = 1
    field_column = 2
    value_column = 7

    values = {}  # {doc_id: {field_name: value}}

    for i in range(1, num_of_rows):  # the num of rows in metadata file
        # Read info from file
        doc = str(sheet.cell(row=i, column=doc_column).value)
        field = str(sheet.cell(row=i, column=field_column).value)
        value = str(sheet.cell(row=i, column=value_column).value)

        # Add to dictionary
        if doc not in values:
            values[doc] = {}

        values[doc][field] = value

    return values


if __name__ == '__main__':
    main()
