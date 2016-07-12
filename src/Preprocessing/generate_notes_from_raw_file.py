'''
This script reads a file containing multiple blobs delineated by Emily's 1234567890qwertyuiop string and produces
a document with the usrId_documentId as the filename.
'''
import openpyxl
import codecs
import re
import SystemUtilities.Globals as g
import SystemUtilities.Configuration as c

# read blob file into memory
with open(c.RAW_DATA_DIR, "rb") as f:
    data = f.read()

# Read in metadata file (excel), creating dict of {doc_id:patient_id}
wb = openpyxl.load_workbook(c.CAISIS_DIR)
sheets = wb.get_sheet_names()
mrn_caisis_dict = dict()
sheet_name = sheets[0]
sheet = wb.get_sheet_by_name(sheet_name)
for i in range(1, g.CAISIS_ROWS, 1):  # the num of rows in metadata file
    mrn_caisis_dict[str(sheet.cell(row=i, column=1).value)] = str(sheet.cell(row=i, column=2).value)

# split on the defined delimiter
blobs = re.split(g.DELIMITER, data)

i = 1
for blob in blobs:
    # split Metadata from the text. \t delimited.
    metadata_and_text = re.split(r"\t", blob)

    if len(metadata_and_text) > g.BLOB_FIELDS:
        parent_id= metadata_and_text[0]
        clinical_event_id = metadata_and_text[1]
        event_date= metadata_and_text[2]
        event_description = metadata_and_text[3]
        addtnl_event_description= metadata_and_text[4]
        doc_text=""
        # Text sometimes spans more than one \t-delimited partition. Take it all, except the last partition, which is the MRN
        for idx in range(8,len(metadata_and_text)-3,1):
            doc_text += metadata_and_text[idx].lstrip().rstrip() + "\n"
        mrn = metadata_and_text[len(metadata_and_text)-2]

        if mrn_caisis_dict.has_key(mrn):
            first_id = mrn_caisis_dict[mrn]
        else:
            first_id = mrn

        #if the id is/contains a 10 digit number AND the first_id is vaguely numeric and not crazy junk
        if re.match(r"\D*(\d{10})\D*", clinical_event_id):
            with open(c.NOTE_OUTPUT_DIR + first_id + "_" + clinical_event_id, "w") as writefile:
                for line in re.split(r"\n", doc_text):
                    writefile.write(line + "\n")
